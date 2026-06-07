from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

orange_fill = PatternFill("solid", fgColor="FFA500")
blue_fill = PatternFill("solid", fgColor="4472C4")
red_fill = PatternFill("solid", fgColor="FF0000")
green_fill = PatternFill("solid", fgColor="70AD47")
header_font = Font(bold=True, color="FFFFFF", name="Arial")


def export(
    df_gefen: pd.DataFrame,
    df_finance: pd.DataFrame | None,
    in_finance_not_gefen: pd.DataFrame | None,
    in_gefen_not_finance: pd.DataFrame | None,
    output_path: str,
    finance_label: str | None = "כספים",
    in_gefen_rejected: pd.DataFrame | None = None,
    in_gefen_no_pdf: pd.DataFrame | None = None,
    gefen_only: bool = False,
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    if gefen_only:
        if in_gefen_rejected is not None:
            _add_result_sheet(wb, "אסמכתאות שנדחו", in_gefen_rejected, red_fill)
        if in_gefen_no_pdf is not None:
            _add_result_sheet(wb, "אסמכתאות ללא סריקה", in_gefen_no_pdf, red_fill)
    else:
        label = finance_label or "כספים"
        _add_result_sheet(wb, f"קיים ב{label} אך לא בגפן", in_finance_not_gefen, red_fill)
        _add_result_sheet(wb, f"משויך בגפן אך לא ב{label}", in_gefen_not_finance, red_fill)
        if in_gefen_rejected is not None:
            _add_result_sheet(wb, "אסמכתאות שנדחו", in_gefen_rejected, red_fill)
        if in_gefen_no_pdf is not None:
            _add_result_sheet(wb, "אסמכתאות ללא סריקה", in_gefen_no_pdf, red_fill)

    wb.save(output_path)
    return output_path


def _add_sheet(wb: Workbook, title: str, df: pd.DataFrame, fill: PatternFill) -> None:
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True
    _write_df(ws, df, fill)


def _add_result_sheet(wb: Workbook, title: str, df: pd.DataFrame, fill: PatternFill) -> None:
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True

    if df.empty:
        ws.cell(row=1, column=1, value="✓ אין פערים").fill = green_fill
        ws.cell(row=1, column=1).font = header_font
    else:
        _write_df(ws, df, fill)


def _write_df(ws, df: pd.DataFrame, fill: PatternFill) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = header_font

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
