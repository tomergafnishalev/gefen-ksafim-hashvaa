"""
Excel and PDF export for tikhnun (budget planning) analysis sections.
Each section can be exported independently.
"""

from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_BLUE_HDR = PatternFill("solid", fgColor="2D5FA0")
_GRAY_ROW = PatternFill("solid", fgColor="E8EDF5")
_RED_FONT = Font(color="C0392B", name="Arial", bold=True)
_WHT_BOLD = Font(color="FFFFFF", name="Arial", bold=True)
_NRM_FONT = Font(name="Arial")
_NRM_BOLD = Font(name="Arial", bold=True)
_MONEY_FMT = "#,##0"
_PCT_FMT = "0%"
_PCT2_FMT = "0.00%"
_ALIGN_R = Alignment(horizontal="right", vertical="center")
_ALIGN_C = Alignment(horizontal="center", vertical="center")


def _rtl(ws):
    ws.sheet_view.rightToLeft = True


def _hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _BLUE_HDR
    c.font = _WHT_BOLD
    c.alignment = _ALIGN_C
    return c


def _col(ws, idx, width):
    ws.column_dimensions[get_column_letter(idx)].width = width


def export_tikhnun_excel(tikhnun: dict, section: str, multiplier: str = "03") -> bytes:
    """
    Export a single tikhnun section to Excel bytes.
    section: "sikar" | "kvua" | "partial" | "yozma"
    multiplier: "03" | "04" (used only for yozma section)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    _rtl(ws)

    if section == "sikar":
        _write_sikar(ws, tikhnun)
        ws.title = "סקירה - גפן"
    elif section == "kvua":
        _write_kvua(ws, tikhnun)
        ws.title = "מימוש תקציב קבוע"
    elif section == "partial":
        _write_partial(ws, tikhnun)
        ws.title = "תוכניות עם ביצוע חלקי"
    elif section == "yozma":
        yozma_key = "yozma_04" if multiplier == "04" else "yozma_03"
        _write_yozma(ws, tikhnun, yozma_key)
        ws.title = "יוזמות וצרכים"
    else:
        raise ValueError(f"Unknown section: {section}")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt_num(v):
    if v is None:
        return ""
    try:
        return int(round(float(v)))
    except Exception:
        return v


def _write_sikar(ws, tikhnun: dict):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    ov = tikhnun.get("overview", {})
    has_doch = tikhnun.get("has_doch", False)

    rows = [
        ("שם מוסד", tikhnun.get("school_name", "")),
        ("סמל מוסד", tikhnun.get("school_code", "")),
        ("שלב מוסד", tikhnun.get("school_stage", "")),
        None,
        ("תקציב גפן", _fmt_num(ov.get("budget"))),
        ("סכום שתוכנן", _fmt_num(ov.get("planned"))),
        ("תקציב קבוע שנותר לתכנון", _fmt_num(ov.get("fixed_gap_abs"))),
        ("תקציב גמיש שנותר לתכנון", _fmt_num(ov.get("flexible_remaining"))),
        None,
        ("סכום חייב בדיווח", _fmt_num(ov.get("sum_chayav"))),
        ("סכום שדווח", _fmt_num(ov.get("sum_divuach"))),
        ("אחוז דיווח (כללי)", ov.get("pct_divuach")),
    ]
    if has_doch and ov.get("pct_tanuz") is not None:
        rows.append(("אחוז דיווח למודל תמרוץ", ov.get("pct_tanuz")))

    ri = 1
    for item in rows:
        if item is None:
            ri += 1
            continue
        label, val = item
        lc = ws.cell(row=ri, column=1, value=label)
        lc.font = _NRM_BOLD
        lc.alignment = _ALIGN_R
        vc = ws.cell(row=ri, column=2, value=val)
        vc.alignment = _ALIGN_R
        if label == "סמל מוסד":
            vc.number_format = "0"
        elif label in ("אחוז דיווח (כללי)", "אחוז דיווח למודל תמרוץ"):
            vc.number_format = _PCT2_FMT
        elif isinstance(val, (int, float)) and val > 100:
            vc.number_format = _MONEY_FMT
        if isinstance(val, (int, float)) and val < 0:
            vc.font = _RED_FONT
        ri += 1


def _write_kvua(ws, tikhnun: dict):
    kvua_rows = tikhnun.get("kvua_rows", [])
    has_multi = tikhnun.get("has_multiple_budget_types", False)

    if has_multi:
        headers = ["סוג תקציב", "שלב חינוך", "סל", "תת סל", "תקציב קבוע", "תקציב שתוכנן", "הפרש שלא תוכנן"]
        widths  = [20, 16, 28, 22, 16, 16, 18]
    else:
        headers = ["שלב חינוך", "סל", "תת סל", "תקציב קבוע", "תקציב שתוכנן", "הפרש שלא תוכנן"]
        widths  = [16, 28, 22, 16, 16, 18]

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    for ci, w in enumerate(widths, 1):
        _col(ws, ci, w)

    for ri, row in enumerate(kvua_rows, 2):
        ci = 1
        if has_multi:
            ws.cell(ri, ci, row.get("budget_type", "")).alignment = _ALIGN_R; ci += 1
        ws.cell(ri, ci, row.get("stage", "")).alignment = _ALIGN_R; ci += 1
        ws.cell(ri, ci, row.get("sal", "")).alignment = _ALIGN_R; ci += 1
        ws.cell(ri, ci, row.get("tatsub", "")).alignment = _ALIGN_R; ci += 1
        c = ws.cell(ri, ci, _fmt_num(row.get("kvua"))); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R; ci += 1
        c = ws.cell(ri, ci, _fmt_num(row.get("tikhnun"))); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R; ci += 1
        h_val = row.get("hefresh", 0)
        c = ws.cell(ri, ci, _fmt_num(h_val)); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
        if h_val < 0:
            c.font = _RED_FONT

    total_ri = len(kvua_rows) + 2
    for ci in range(1, len(headers) + 1):
        ws.cell(total_ri, ci).fill = _GRAY_ROW

    ci = 1
    if has_multi:
        ci += 1
    ws.cell(total_ri, ci, 'סה"כ').font = _NRM_BOLD
    ws.cell(total_ri, ci).alignment = _ALIGN_R
    ci += 2  # skip sal, go to kvua
    c = ws.cell(total_ri, ci, _fmt_num(sum(r.get("kvua", 0) for r in kvua_rows)))
    c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R; c.font = _NRM_BOLD; ci += 1
    c = ws.cell(total_ri, ci, _fmt_num(sum(r.get("tikhnun", 0) for r in kvua_rows)))
    c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R; c.font = _NRM_BOLD; ci += 1
    total_h = sum(r.get("hefresh", 0) for r in kvua_rows)
    c = ws.cell(total_ri, ci, _fmt_num(total_h))
    c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
    c.font = _RED_FONT if total_h < 0 else _NRM_BOLD


def _write_partial(ws, tikhnun: dict):
    partial_rows = tikhnun.get("partial_rows", [])
    headers = ["קוד דיווח", "שם מענה", "מספר מענה", "תכנון", "דיווח", "הפרש", "אחוז דיווח תוכנית"]
    widths  = [12, 38, 14, 14, 14, 14, 18]

    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)
    for ci, w in enumerate(widths, 1):
        _col(ws, ci, w)

    for ri, row in enumerate(partial_rows, 2):
        ws.cell(ri, 1, row.get("rcode", "")).alignment = _ALIGN_R
        ws.cell(ri, 2, row.get("name", "")).alignment = _ALIGN_R
        ws.cell(ri, 3, row.get("mispnum", "")).alignment = _ALIGN_R
        c = ws.cell(ri, 4, _fmt_num(row.get("tikhnun"))); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
        c = ws.cell(ri, 5, _fmt_num(row.get("divuach"))); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
        h_val = row.get("hefresh", 0)
        c = ws.cell(ri, 6, _fmt_num(h_val)); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
        if h_val < 0:
            c.font = _RED_FONT
        pct = row.get("pct", 0)
        c = ws.cell(ri, 7, pct); c.number_format = _PCT2_FMT; c.alignment = _ALIGN_R

    total_ri = len(partial_rows) + 2
    for ci in range(1, len(headers) + 1):
        ws.cell(total_ri, ci).fill = _GRAY_ROW
    ws.cell(total_ri, 1, 'סה"כ הפרש לטיפול').font = _NRM_BOLD
    ws.cell(total_ri, 1).alignment = _ALIGN_R
    total_h = tikhnun.get("sum_hefresh_partial", 0)
    c = ws.cell(total_ri, 6, _fmt_num(total_h))
    c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R; c.font = _NRM_BOLD


def _write_yozma(ws, tikhnun: dict, yozma_key: str):
    yozma = tikhnun.get(yozma_key, tikhnun.get("yozma_03", {}))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22

    title = ws.cell(1, 1, "יוזמות וצרכים ייחודיים")
    title.font = Font(name="Arial", bold=True, size=12)
    title.alignment = _ALIGN_R

    ws.cell(3, 1, "תקציב מקסימלי לתכנון יוזמות").font = _NRM_BOLD
    ws.cell(3, 1).alignment = _ALIGN_R
    c = ws.cell(3, 2, _fmt_num(yozma.get("max")))
    c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R

    ws.cell(4, 1, "בתכנון").font = _NRM_BOLD
    ws.cell(4, 1).alignment = _ALIGN_R
    c = ws.cell(4, 2, _fmt_num(yozma.get("betikhnun")))
    c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R

    hefresh_val = yozma.get("hefresh", 0)
    ws.cell(5, 1, "הפרש").font = _NRM_BOLD
    ws.cell(5, 1).alignment = _ALIGN_R
    c = ws.cell(5, 2, _fmt_num(hefresh_val))
    c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
    if hefresh_val < 0:
        c.font = _RED_FONT

    for ci, h in enumerate(["סעיף", "תקרה", "בתכנון", "הפרש שניתן לתכנון"], 1):
        _hdr(ws, 7, ci, h)

    for ri, item in enumerate(yozma.get("detail", []), 8):
        ws.cell(ri, 1, item.get("label", "")).alignment = _ALIGN_R
        c = ws.cell(ri, 2, _fmt_num(item.get("cap"))); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
        c = ws.cell(ri, 3, _fmt_num(item.get("betikhnun"))); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
        h_val = item.get("hefresh", 0)
        c = ws.cell(ri, 4, _fmt_num(h_val)); c.number_format = _MONEY_FMT; c.alignment = _ALIGN_R
        if h_val < 0:
            c.font = _RED_FONT


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def build_tikhnun_section_story(tikhnun: dict, section: str, multiplier: str = "03") -> list:
    """Return ReportLab flowables for a single tikhnun section (used by combined exporter)."""
    from bidi.algorithm import get_display
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from logic.pdf_exporter import _ensure_fonts, _FONT_NAME, _FONT_BOLD

    _ensure_fonts()
    font_name = _FONT_NAME
    font_bold = _FONT_BOLD

    def rtl(text):
        if not text: return ""
        s = str(text)
        return get_display(s) if any("֐" <= c <= "׿" for c in s) else s

    def fmt_money(v):
        if v is None: return ""
        try: return f"{int(round(float(v))):,}"
        except: return str(v)

    def fmt_pct(v):
        if v is None: return ""
        try: return f"{float(v)*100:.2f}%"
        except: return str(v)

    HDR_COLOR = colors.HexColor("#2D5FA0")
    GRAY = colors.HexColor("#E8EDF5")

    def _tbl_style(num_cols, num_rows, total_row=None):
        style = [
            ("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), HDR_COLOR), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), font_bold), ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ]
        if total_row is not None:
            style += [("BACKGROUND", (0, total_row), (-1, total_row), GRAY),
                      ("FONTNAME", (0, total_row), (-1, total_row), font_bold)]
        return TableStyle(style)

    style_title = ParagraphStyle("tkh_title", fontName=font_bold, fontSize=13, alignment=1)

    def _rev(data): return [list(reversed(row)) for row in data]

    story = []

    if section == "sikar":
        story.append(Paragraph(rtl("סקירה - גפן"), style_title))
        story.append(Spacer(1, 0.4*cm))
        ov = tikhnun.get("overview", {})
        has_doch = tikhnun.get("has_doch", False)
        data = [
            [rtl("שדה"), rtl("ערך")],
            [rtl("שם מוסד"), rtl(tikhnun.get("school_name", ""))],
            [rtl("סמל מוסד"), str(tikhnun.get("school_code", ""))],
            [rtl("שלב מוסד"), rtl(tikhnun.get("school_stage", ""))],
            [rtl("תקציב גפן"), fmt_money(ov.get("budget"))],
            [rtl("סכום שתוכנן"), fmt_money(ov.get("planned"))],
            [rtl("תקציב קבוע שנותר"), fmt_money(ov.get("fixed_gap_abs"))],
            [rtl("תקציב גמיש שנותר"), fmt_money(ov.get("flexible_remaining"))],
            [rtl("סכום חייב בדיווח"), fmt_money(ov.get("sum_chayav"))],
            [rtl("סכום שדווח"), fmt_money(ov.get("sum_divuach"))],
            [rtl("אחוז דיווח (כללי)"), fmt_pct(ov.get("pct_divuach"))],
        ]
        if has_doch and ov.get("pct_tanuz") is not None:
            data.append([rtl("אחוז דיווח למודל תמרוץ"), fmt_pct(ov.get("pct_tanuz"))])
        data = _rev(data)
        tbl = Table(data, colWidths=[7*cm, 9*cm])
        tbl.setStyle(_tbl_style(2, len(data)))
        story.append(tbl)

    elif section == "kvua":
        story.append(Paragraph(rtl("מימוש תקציב קבוע - גפן"), style_title))
        story.append(Spacer(1, 0.4*cm))
        kvua_rows = tikhnun.get("kvua_rows", [])
        has_multi = tikhnun.get("has_multiple_budget_types", False)
        if has_multi:
            headers = ["סוג תקציב", "שלב", "סל", "תת סל", "תקציב קבוע", "תוכנן", "הפרש"]
            col_widths = [3*cm, 2.5*cm, 3.5*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm]
        else:
            headers = ["שלב", "סל", "תת סל", "תקציב קבוע", "תוכנן", "הפרש"]
            col_widths = [2.5*cm, 4*cm, 3.5*cm, 2.8*cm, 2.8*cm, 2.8*cm]
        data = [[rtl(h) for h in headers]]
        for row in kvua_rows:
            r = []
            if has_multi: r.append(rtl(row.get("budget_type", "")))
            r += [rtl(row.get("stage", "")), rtl(row.get("sal", "")), rtl(row.get("tatsub", "")),
                  fmt_money(row.get("kvua")), fmt_money(row.get("tikhnun")), fmt_money(row.get("hefresh"))]
            data.append(r)
        total_h = sum(r.get("hefresh", 0) for r in kvua_rows)
        totals = [""] * (len(headers) - 3) + [rtl('סה"כ'), "",
                  fmt_money(sum(r.get("kvua", 0) for r in kvua_rows)),
                  fmt_money(sum(r.get("tikhnun", 0) for r in kvua_rows)), fmt_money(total_h)]
        data.append(totals)
        data = _rev(data)
        tbl = Table(data, colWidths=list(reversed(col_widths)))
        tbl.setStyle(_tbl_style(len(headers), len(data), total_row=len(data)-1))
        story.append(tbl)

    elif section == "partial":
        story.append(Paragraph(rtl("תוכניות עם ביצוע חלקי"), style_title))
        story.append(Spacer(1, 0.4*cm))
        partial_rows = tikhnun.get("partial_rows", [])
        if not partial_rows:
            story.append(Paragraph(rtl("אין תוכניות עם ביצוע חלקי"),
                                   ParagraphStyle("tkh_hdr", fontName=font_bold, fontSize=9)))
        else:
            headers = ["קוד", "שם מענה", "מס' מענה", "תכנון", "דיווח", "הפרש", "אחוז"]
            col_widths = [1.5*cm, 6*cm, 2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 1.8*cm]
            data = [[rtl(h) for h in headers]]
            for row in partial_rows:
                data.append([rtl(row.get("rcode", "")), rtl(row.get("name", "")),
                              rtl(row.get("mispnum", "")), fmt_money(row.get("tikhnun")),
                              fmt_money(row.get("divuach")), fmt_money(row.get("hefresh")),
                              fmt_pct(row.get("pct"))])
            total_h = tikhnun.get("sum_hefresh_partial", 0)
            data.append([rtl('סה"כ הפרש'), "", "", "", "", fmt_money(total_h), ""])
            data = _rev(data)
            tbl = Table(data, colWidths=list(reversed(col_widths)))
            tbl.setStyle(_tbl_style(len(headers), len(data), total_row=len(data)-1))
            story.append(tbl)

    elif section == "yozma":
        story.append(Paragraph(rtl("יוזמות וצרכים - פירוט"), style_title))
        story.append(Spacer(1, 0.4*cm))
        yozma_key = "yozma_04" if multiplier == "04" else "yozma_03"
        yozma = tikhnun.get(yozma_key, tikhnun.get("yozma_03", {}))
        summary_data = [
            [rtl("תקציב מקסימלי"), fmt_money(yozma.get("max"))],
            [rtl("בתכנון"), fmt_money(yozma.get("betikhnun"))],
            [rtl("הפרש"), fmt_money(yozma.get("hefresh"))],
        ]
        tbl = Table(_rev(summary_data), colWidths=[7*cm, 9*cm])
        tbl.setStyle(_tbl_style(2, 3))
        story.append(tbl)
        story.append(Spacer(1, 0.4*cm))
        headers = ["סעיף", "תקרה", "בתכנון", "הפרש שניתן לתכנון"]
        col_widths = [5*cm, 4*cm, 4*cm, 5*cm]
        data = [[rtl(h) for h in headers]]
        for item in yozma.get("detail", []):
            data.append([rtl(item.get("label", "")), fmt_money(item.get("cap")),
                         fmt_money(item.get("betikhnun")), fmt_money(item.get("hefresh"))])
        tbl2 = Table(_rev(data), colWidths=list(reversed(col_widths)))
        tbl2.setStyle(_tbl_style(4, len(data)))
        story.append(tbl2)

    return story


def export_tikhnun_pdf(tikhnun: dict, section: str, multiplier: str = "03") -> bytes:
    """Export a single tikhnun section to PDF bytes using ReportLab."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, PageBreak
    from logic.pdf_exporter import _ensure_fonts, build_school_info_story

    _ensure_fonts()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    story = []
    # Always prepend school info for non-sikar sections (sikar IS the school info)
    if section != "sikar":
        story.extend(build_school_info_story(tikhnun))
        story.append(PageBreak() if section in ("kvua", "partial", "yozma") else __import__("reportlab.platypus", fromlist=["Spacer"]).Spacer(1, 0.3*__import__("reportlab.lib.units", fromlist=["cm"]).cm))

    story.extend(build_tikhnun_section_story(tikhnun, section, multiplier))
    doc.build(story)
    return buf.getvalue()


# Keep the old PDF helpers block for backward-compat (not used anymore but harmless)
def _legacy_pdf_setup():
    """Kept for reference only — logic moved to build_tikhnun_section_story."""
    pass


def _tbl_style_compat(num_cols, num_rows, total_row=None):
    """Shim for internal callers — unused."""
    pass

    def _rev(data):
        """Reverse column order of every row so RTL columns appear correctly in LTR PDF."""
        return [list(reversed(row)) for row in data]

    story = []

    if section == "sikar":
        story.append(Paragraph(rtl("סקירה - גפן"), style_title))
        story.append(Spacer(1, 0.4*cm))
        ov = tikhnun.get("overview", {})
        has_doch = tikhnun.get("has_doch", False)
        data = [
            [rtl("שדה"), rtl("ערך")],
            [rtl("שם מוסד"), rtl(tikhnun.get("school_name", ""))],
            [rtl("סמל מוסד"), str(tikhnun.get("school_code", ""))],
            [rtl("שלב מוסד"), rtl(tikhnun.get("school_stage", ""))],
            [rtl("תקציב גפן"), fmt_money(ov.get("budget"))],
            [rtl("סכום שתוכנן"), fmt_money(ov.get("planned"))],
            [rtl("תקציב קבוע שנותר"), fmt_money(ov.get("fixed_gap_abs"))],
            [rtl("תקציב גמיש שנותר"), fmt_money(ov.get("flexible_remaining"))],
            [rtl("סכום חייב בדיווח"), fmt_money(ov.get("sum_chayav"))],
            [rtl("סכום שדווח"), fmt_money(ov.get("sum_divuach"))],
            [rtl("אחוז דיווח (כללי)"), fmt_pct(ov.get("pct_divuach"))],
        ]
        if has_doch and ov.get("pct_tanuz") is not None:
            data.append([rtl("אחוז דיווח למודל תמרוץ"), fmt_pct(ov.get("pct_tanuz"))])
        data = _rev(data)  # label col rightmost in RTL
        tbl = Table(data, colWidths=[7*cm, 9*cm])
        tbl.setStyle(_tbl_style(2, len(data)))
        story.append(tbl)

    elif section == "kvua":
        story.append(Paragraph(rtl("מימוש תקציב קבוע - גפן"), style_title))
        story.append(Spacer(1, 0.4*cm))
        kvua_rows = tikhnun.get("kvua_rows", [])
        has_multi = tikhnun.get("has_multiple_budget_types", False)
        if has_multi:
            headers = ["סוג תקציב", "שלב", "סל", "תת סל", "תקציב קבוע", "תוכנן", "הפרש"]
            col_widths = [3*cm, 2.5*cm, 3.5*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm]
        else:
            headers = ["שלב", "סל", "תת סל", "תקציב קבוע", "תוכנן", "הפרש"]
            col_widths = [2.5*cm, 4*cm, 3.5*cm, 2.8*cm, 2.8*cm, 2.8*cm]
        data = [[rtl(h) for h in headers]]
        for row in kvua_rows:
            r = []
            if has_multi:
                r.append(rtl(row.get("budget_type", "")))
            r += [rtl(row.get("stage", "")), rtl(row.get("sal", "")), rtl(row.get("tatsub", "")),
                  fmt_money(row.get("kvua")), fmt_money(row.get("tikhnun")), fmt_money(row.get("hefresh"))]
            data.append(r)
        total_h = sum(r.get("hefresh", 0) for r in kvua_rows)
        totals = [""] * (len(headers) - 3) + [rtl('סה"כ'), "", fmt_money(sum(r.get("kvua", 0) for r in kvua_rows)),
                  fmt_money(sum(r.get("tikhnun", 0) for r in kvua_rows)), fmt_money(total_h)]
        data.append(totals)
        data = _rev(data)
        col_widths = list(reversed(col_widths))
        tbl = Table(data, colWidths=col_widths)
        tbl.setStyle(_tbl_style(len(headers), len(data), total_row=len(data)-1))
        story.append(tbl)

    elif section == "partial":
        story.append(Paragraph(rtl("תוכניות עם ביצוע חלקי"), style_title))
        story.append(Spacer(1, 0.4*cm))
        partial_rows = tikhnun.get("partial_rows", [])
        if not partial_rows:
            story.append(Paragraph(rtl("אין תוכניות עם ביצוע חלקי"), style_hdr))
        else:
            headers = ["קוד", "שם מענה", "מס' מענה", "תכנון", "דיווח", "הפרש", "אחוז"]
            col_widths = [1.5*cm, 6*cm, 2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 1.8*cm]
            data = [[rtl(h) for h in headers]]
            for row in partial_rows:
                data.append([
                    rtl(row.get("rcode", "")),
                    rtl(row.get("name", "")),
                    rtl(row.get("mispnum", "")),
                    fmt_money(row.get("tikhnun")),
                    fmt_money(row.get("divuach")),
                    fmt_money(row.get("hefresh")),
                    fmt_pct(row.get("pct")),
                ])
            total_h = tikhnun.get("sum_hefresh_partial", 0)
            data.append([rtl('סה"כ הפרש'), "", "", "", "", fmt_money(total_h), ""])
            data = _rev(data)
            col_widths = list(reversed(col_widths))
            tbl = Table(data, colWidths=col_widths)
            tbl.setStyle(_tbl_style(len(headers), len(data), total_row=len(data)-1))
            story.append(tbl)

    elif section == "yozma":
        story.append(Paragraph(rtl("יוזמות וצרכים - פירוט"), style_title))
        story.append(Spacer(1, 0.4*cm))
        yozma_key = "yozma_04" if multiplier == "04" else "yozma_03"
        yozma = tikhnun.get(yozma_key, tikhnun.get("yozma_03", {}))
        summary_data = [
            [rtl("תקציב מקסימלי"), fmt_money(yozma.get("max"))],
            [rtl("בתכנון"), fmt_money(yozma.get("betikhnun"))],
            [rtl("הפרש"), fmt_money(yozma.get("hefresh"))],
        ]
        summary_data = _rev(summary_data)
        tbl = Table(summary_data, colWidths=[7*cm, 9*cm])
        tbl.setStyle(_tbl_style(2, 3))
        story.append(tbl)
        story.append(Spacer(1, 0.4*cm))
        headers = ["סעיף", "תקרה", "בתכנון", "הפרש שניתן לתכנון"]
        col_widths = [5*cm, 4*cm, 4*cm, 5*cm]
        data = [[rtl(h) for h in headers]]
        for item in yozma.get("detail", []):
            data.append([rtl(item.get("label", "")),
                         fmt_money(item.get("cap")),
                         fmt_money(item.get("betikhnun")),
                         fmt_money(item.get("hefresh"))])
        data = _rev(data)
        col_widths = list(reversed(col_widths))
        tbl2 = Table(data, colWidths=col_widths)
        tbl2.setStyle(_tbl_style(4, len(data)))
        story.append(tbl2)

    doc.build(story)
    return buf.getvalue()
