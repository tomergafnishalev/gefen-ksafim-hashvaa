from pathlib import Path

import openpyxl


def identify_file(filename: str) -> str:
    """Return 'gefen', 'kesafim2000', 'payscool', or 'unknown'."""
    ext = Path(filename).suffix.lower()

    if ext == ".xls":
        return _identify_xls(filename)

    if ext == ".xlsx":
        return _identify_xlsx(filename)

    return "unknown"


def _identify_xls(filename: str) -> str:
    """XLS files are either Kesafim2000 (TSV disguised as XLS) or unknown."""
    try:
        with open(filename, "r", encoding="iso-8859-8") as f:
            first_line = f.readline().rstrip("\r\n")
        parts = first_line.split("\t")
        a1 = parts[0].strip() if len(parts) > 0 else ""
        d1 = parts[3].strip() if len(parts) > 3 else ""
        if a1 == "קוד גפן" and d1 == "סוג תקציב":
            return "kesafim2000"
    except Exception:
        pass
    return "unknown"


def _identify_xlsx(filename: str) -> str:
    """Identify XLSX files by sheet name + cell content."""
    try:
        wb = openpyxl.load_workbook(filename, read_only=True)
    except Exception:
        return "unknown"

    sheets = wb.sheetnames

    # --- Gefen ---
    if "דיווח ביצוע" in sheets:
        ws = wb["דיווח ביצוע"]
        row1 = _get_row(ws, 1)
        if (
            _cell(row1, 0) == "מסלול רכישה"
            and _cell(row1, 2) == "סוג מענה"
            and _cell(row1, 3) == "שם מענה"
            and _cell(row1, 13) == "האם קיים קובץ"
        ):
            wb.close()
            return "gefen"

    # --- PaySchool --- (scan all sheets, sheet name may vary)
    # Supports both old format (סעיף at col 0) and new format (extra cols before סעיף)
    for sheet_name in sheets:
        ws = wb[sheet_name]
        row4 = _get_row(ws, 4)
        row4_vals = {str(v).strip() for v in row4 if v is not None}
        if (
            "סעיף" in row4_vals
            and "ח.פ" in row4_vals
            and "סוג חשבונית" in row4_vals
        ):
            wb.close()
            return "payscool"

    # --- Kesafim2000 saved as XLSX ---
    for sheet_name in sheets:
        ws = wb[sheet_name]
        row1 = _get_row(ws, 1)
        if _cell(row1, 0) == "קוד גפן" and _cell(row1, 3) == "סוג תקציב":
            wb.close()
            return "kesafim2000"

    # --- SchoolCash ---
    ws = wb[sheets[0]]
    row1 = _get_row(ws, 1)
    if (
        _cell(row1, 5) == "עוסק מורשה"
        and _cell(row1, 6) == "מספר תעודה"
        and _cell(row1, 7) == "סוג תעודה"
        and _cell(row1, 14) == "תאור שורה בחשבונית"
    ):
        wb.close()
        return "schoolcash"

    # --- Tikhnun (budget planning) ---
    # Exactly 2 sheets; one has K1="השתתפות רשות/ בעלות" + O1="תאריך אחרון לאישור רשות";
    # the other has O1="עלות מענה כוללת" + P1="עלות מתקציב"
    if len(sheets) == 2:
        hakol_found = False
        perut_found = False
        for sheet_name in sheets:
            ws2 = wb[sheet_name]
            r1 = _get_row(ws2, 1)
            if _cell(r1, 10) == "השתתפות רשות/ בעלות" and _cell(r1, 14) == "תאריך אחרון לאישור רשות":
                hakol_found = True
            if _cell(r1, 14) == "עלות מענה כוללת" and _cell(r1, 15) == "עלות מתקציב":
                perut_found = True
        if hakol_found and perut_found:
            wb.close()
            return "tikhnun"

    wb.close()
    return "unknown"


def _get_row(ws, row_num: int) -> list:
    rows = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    return list(rows[0]) if rows else []


def _cell(row: list, index: int) -> str:
    if index < len(row) and row[index] is not None:
        return str(row[index]).strip()
    return ""
