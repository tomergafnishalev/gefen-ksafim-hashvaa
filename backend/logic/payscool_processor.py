import re

import pandas as pd

from logic.gefen_processor import normalize_amount


def load_payscool(filepath: str) -> tuple[pd.DataFrame, int]:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet_name = None
    for s in wb.sheetnames:
        ws = wb[s]
        rows = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        if rows and "סעיף" in {str(v).strip() for v in rows[0] if v is not None}:
            sheet_name = s
            break
    wb.close()
    df = pd.read_excel(filepath, sheet_name=sheet_name or 0, header=None)
    df.columns = df.iloc[3]
    df = df.iloc[4:].reset_index(drop=True)

    df["report_code"] = df["סעיף"].apply(_extract_report_code)
    df = df[df["report_code"].notna()].copy()

    cancelled_count = int((df["סטטוס חשבונית"] == "מבוטלת").sum())
    df = df[df["סטטוס חשבונית"] != "מבוטלת"].copy()

    df["amount"] = df['סה"כ לסעיף'].apply(normalize_amount)
    df["ichud"] = (
        df["ח.פ"].apply(normalize_amount)
        + "-"
        + df["מספר חשבונית"].apply(normalize_amount)
        + "-"
        + df["report_code"].astype(str)
        + "-"
        + df["amount"]
    )
    return df, cancelled_count


def _extract_report_code(value) -> str | None:
    match = re.search(r"\((\d+)\)", str(value))
    return match.group(1) if match else None
