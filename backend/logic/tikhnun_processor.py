"""
Tikhnun (budget planning) processor.
Reads a 2-sheet budget planning file ("הכל" + "פירוט המענים") and optionally
cross-references with a Gefen doch (execution report) to produce analysis data.
"""

import re
import datetime
import openpyxl
from collections import defaultdict


STAGE_MAP = {
    "עליונה בלבד": "תיכון",
    'חט"ב בלבד': "חטיבת ביניים",
    "יסודי בלבד": "חטיבת ביניים",
    "יסודי וחט\"ב": "חטיבת ביניים",
    "יסודי": "חטיבת ביניים",
}

SAL_ORDER = [
    "מענים פדגוגיים ורגשיים",
    "חינוך חברתי - קהילתי והעשרה",
    "אוכלוסיות במיקוד",
    "קידום רווחת התלמיד",
    "תשתיות בית ספריות",
]

# Yozma codes by stage keyword
YOZMA_CODES = {
    "תיכון": {"106": "נלוות", "107": "רכוש קבוע", "105": "כיבוד", "108": "תיקונים קלים"},
    "חטיבת ביניים": {"69": "נלוות", "70": "רכוש קבוע", "68": "כיבוד", "71": "תיקונים קלים"},
}


def _to_num(v):
    if v is None or str(v).strip() == "":
        return 0
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0


def _fmt_pct(v):
    try:
        s = str(v).replace("%", "").strip()
        f = float(s)
        return f / 100 if f > 1 else f
    except Exception:
        return 0


def _sal_order_key(s):
    for i, k in enumerate(SAL_ORDER):
        if k in s:
            return i
    return 99


def _extract_plan_num(name):
    if not name:
        return ""
    s = str(name).strip()
    left = s.split("-")[0].strip()
    if left.isdigit():
        return left
    return ""


def _fmt_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if not s or s in ("nan", "None"):
        return ""
    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", s):
        d, m, y = s.split("/")
        return f"{int(d):02d}/{int(m):02d}/{y}"
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        y, m, d = s[:10].split("-")
        return f"{int(d):02d}/{int(m):02d}/{y}"
    return s


def _plan_key_perut(r):
    j = str(r[9]).strip() if r[9] else ""
    if j and j != " ":
        return j
    rcode = str(r[17]).strip() if r[17] else ""
    iname = str(r[8]).strip() if r[8] else ""
    return f"{rcode}-{iname}"


def _plan_key_doch(r):
    d = str(r[3]).strip() if r[3] else ""
    b = str(r[1]).strip() if r[1] else ""
    pnum = _extract_plan_num(d)
    if pnum:
        return pnum
    return f"{b}-{d}"


def _norm_quotes(s: str) -> str:
    return str(s).replace("״", '"').replace("“", '"').replace("”", '"')


def normalize_budget_name(raw: str) -> str:
    raw = _norm_quotes(raw).strip()
    if 'גפ"ן' in raw or "גפן" in raw.replace('"', ""):
        return "גפן חירום" if "חירום" in raw else "גפן"
    if "תנופה" in raw:
        return "תנופה"
    if "דוקאטי" in raw:
        return "דוקאטי"
    if 'פל"ג' in raw or "פלג" in raw:
        return 'פל"ג'
    return raw


def _build_yozma_detail(yozma_codes, yozma_by_code, total_betikhnun, max_yozma, gamish_notar):
    codes = list(yozma_codes.keys())
    caps = {
        codes[0]: max_yozma,
        codes[1]: max_yozma * 0.5,
        codes[2]: max_yozma * 0.15,
        codes[3]: max_yozma * 0.1,
    }
    detail = []
    for code, label in yozma_codes.items():
        betikhnun = yozma_by_code.get(code, 0)
        cap = caps[code]
        diff = cap - betikhnun
        hefresh_shanim = min(diff, gamish_notar) if diff >= 0 else diff
        detail.append({
            "label": label,
            "code": code,
            "cap": round(cap),
            "betikhnun": round(betikhnun),
            "hefresh": round(hefresh_shanim),
        })
    hefresh_total = round(max_yozma) - round(total_betikhnun)
    return {
        "max": round(max_yozma),
        "betikhnun": round(total_betikhnun),
        "hefresh": hefresh_total,
        "is_negative": hefresh_total < 0,
        "detail": detail,
    }


def load_tikhnun(filepath):
    """
    Load and parse a tikhnun budget planning file.
    Returns a dict with all parsed data, or raises ValueError if file is invalid.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheets = wb.sheetnames
        if len(sheets) != 2:
            wb.close()
            raise ValueError(f"Expected 2 sheets, got {len(sheets)}")

        hakol_ws = None
        perut_ws = None
        for sh in sheets:
            ws = wb[sh]
            rows_peek = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            if not rows_peek:
                continue
            header = rows_peek[0]
            if len(header) > 14 and header[10] == "השתתפות רשות/ בעלות" and header[14] == "תאריך אחרון לאישור רשות":
                hakol_ws = sh
            if len(header) > 15 and header[14] == "עלות מענה כוללת" and header[15] == "עלות מתקציב":
                perut_ws = sh

        if not hakol_ws or not perut_ws:
            wb.close()
            raise ValueError("Could not identify הכל / פירוט המענים sheets")

        hakol_rows = [list(r) for r in wb[hakol_ws].iter_rows(values_only=True)]
        perut_rows = [list(r) for r in wb[perut_ws].iter_rows(values_only=True)]
        wb.close()
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Failed to read tikhnun file: {e}")

    return _process_tikhnun(hakol_rows, perut_rows)


def _process_tikhnun(hakol_rows, perut_rows):
    hakol_data = hakol_rows[1:]  # skip header

    # School info from first data row
    school_name = str(hakol_data[0][3]).strip() if hakol_data[0][3] else ""
    school_code = hakol_data[0][2]
    school_stage_raw = str(hakol_data[0][4]).strip() if hakol_data[0][4] else ""
    school_stage = STAGE_MAP.get(school_stage_raw, school_stage_raw)

    # Control row: budget type contains גפ"ן AND F = תקציב כללי
    ctrl = None
    machozi_I = 0
    for r in hakol_data:
        aval = _norm_quotes(str(r[0]).strip()) if r[0] else ""
        fval = str(r[5]).strip() if r[5] else ""
        if ('גפ"ן' in aval or "גפן" in aval.replace('"', "")):
            if "כללי" in fval and ctrl is None:
                ctrl = r
            if "מחוזי" in fval and "דיפרנציאלי" in fval:
                machozi_I = _to_num(r[8])

    if ctrl is None:
        # No גפן ctrl row — fall back to first non-summary budget row with H > 0
        for r in hakol_data:
            aval = _norm_quotes(str(r[0]).strip()) if r[0] else ""
            if "סה" in aval[:4]:
                continue
            if aval and _to_num(r[7]) > 0:
                ctrl = r
                break

    if ctrl is None:
        raise ValueError("Could not find control row (תקציב גפ\"ן + תקציב כללי)")

    H = _to_num(ctrl[7])   # total budget
    L = _to_num(ctrl[11])  # planned
    S = _to_num(ctrl[18])  # reported
    T = _fmt_pct(ctrl[19]) # % reported

    # Fixed budget rows (kvua)
    kvua_rows = []
    for r in hakol_data:
        fval = str(r[5]).strip() if r[5] else ""
        gval = str(r[6]).strip() if r[6] else ""
        i_val = _to_num(r[8])
        l_val = _to_num(r[11])
        budget_type = _norm_quotes(str(r[0]).strip()) if r[0] else ""

        cond1 = "קידום רווחת התלמיד" in fval and i_val > 0
        cond2 = bool(gval) and i_val > 0

        if cond1 or cond2:
            stage = STAGE_MAP.get(str(r[4]).strip(), str(r[4]).strip()) if r[4] else ""
            sal = fval.replace("סל ", "")
            if sal == "קידום רווחת התלמיד":
                sal = "רווחת התלמיד"
            tatsub = gval
            if "STEM" in tatsub:
                tatsub = "STEM"
            plan = min(l_val, i_val)
            diff = plan - i_val
            kvua_rows.append({
                "budget_type": budget_type,
                "stage": stage,
                "sal": sal,
                "tatsub": tatsub,
                "kvua": i_val,
                "tikhnun": plan,
                "hefresh": diff,
            })

    kvua_rows.sort(key=lambda x: (x["stage"] != "תיכון", _sal_order_key(x["sal"])))
    hefresh_kvua_total = sum(r["hefresh"] for r in kvua_rows)
    abs_hefresh_kvua = abs(hefresh_kvua_total)
    budget_types_in_kvua = set(r["budget_type"] for r in kvua_rows)
    has_multiple_budget_types = len(budget_types_in_kvua) > 1

    gamish_notar = H - L - abs_hefresh_kvua
    max_yozma_03 = (H + machozi_I) * 0.3
    max_yozma_04 = (H + machozi_I) * 0.4

    # Unique plans from פירוט המענים
    seen_keys = {}
    for r in perut_rows[1:]:
        key = tuple(str(x).strip() if x else "" for x in r[:10])
        if key not in seen_keys:
            seen_keys[key] = r

    unique_plans = list(seen_keys.values())

    # Plans that require reporting (have report code)
    plans_with_R = [r for r in unique_plans if r[17] and str(r[17]).strip()]
    sum_chayav = sum(_to_num(r[15]) for r in plans_with_R)

    # Yozma codes for this stage
    yozma_codes = YOZMA_CODES.get(school_stage, YOZMA_CODES["תיכון"])
    yozma_by_code = defaultdict(float)
    _comp_seen: set = set()
    for r in perut_rows[1:]:
        rcode = str(r[17] or "").strip()
        if rcode not in yozma_codes:
            continue
        _c13 = r[13] if len(r) > 13 else None
        if _c13 is not None:
            comp_key = (
                str(r[9]  or "").strip(),  # J - מספר מענה
                str(r[10] or "").strip(),  # K
                str(r[11] or "").strip(),  # L
                str(r[12] or "").strip(),  # M
                _c13,                       # N - עלות למרכיב
                rcode,                      # R - קוד דיווח
            )
        else:
            comp_key = (tuple(str(x).strip() if x else "" for x in r[:10]), rcode)
        if comp_key in _comp_seen:
            continue
        _comp_seen.add(comp_key)
        yozma_by_code[rcode] += _to_num(_c13) if _c13 is not None else _to_num(r[15])

    total_yozma_betikhnun = sum(yozma_by_code.values())

    yozma_03 = _build_yozma_detail(yozma_codes, dict(yozma_by_code), total_yozma_betikhnun, max_yozma_03, gamish_notar)
    yozma_04 = _build_yozma_detail(yozma_codes, dict(yozma_by_code), total_yozma_betikhnun, max_yozma_04, gamish_notar)

    # ── Multi-budget detection ────────────────────────────────────────────────
    seen_budget_names: set = set()
    _budgets_data: list = []
    for r in hakol_data:
        aval = _norm_quotes(str(r[0]).strip()) if r[0] else ""
        if "סה" in aval[:4]:   # "סה"כ..." summary rows — skip
            continue
        H_b = _to_num(r[7])
        if aval and H_b > 0 and aval not in seen_budget_names:
            seen_budget_names.add(aval)
            _budgets_data.append({
                "raw_name": aval,
                "norm_name": normalize_budget_name(aval),
                "H": H_b,
                "L": _to_num(r[11]),
                "S": _to_num(r[18]),
                "T": _fmt_pct(r[19]),
                "machozi_I": 0,
                "yozma_breakdown": [],
                "nihul_breakdown": [],
            })
        elif aval in seen_budget_names:
            break

    # Deduplicate by normalized name — keep only first occurrence per norm_name.
    # Two raw budget names can normalize to "גפן" (e.g. "גפ"ן הכללי" and "גפ"ן מחוזי");
    # keeping both causes duplicate pills and mis-distributed breakdown items.
    seen_nn: set = set()
    deduped: list = []
    for b in _budgets_data:
        if b["norm_name"] not in seen_nn:
            seen_nn.add(b["norm_name"])
            deduped.append(b)
    _budgets_data = deduped

    # Normalized name of the control (main גפן) budget row — used for gamish_notar
    ctrl_norm = normalize_budget_name(_norm_quotes(str(ctrl[0]).strip())) if ctrl[0] else ""

    # The ctrl row ("כללי") is the authoritative source for the גפן budget's H/L/S/T.
    # The first "גפן" row in hakol may be a combined summary row with a larger H.
    # Override the גפן entry with ctrl values to keep per-budget max consistent with global.
    for b in _budgets_data:
        if b["norm_name"] == ctrl_norm:
            b["H"] = H   # ctrl row's H (identical to global H used for max_yozma)
            b["L"] = L
            b["S"] = S
            b["T"] = T
            break

    # machozi per budget (second pass — need full scan)
    for r in hakol_data:
        aval = _norm_quotes(str(r[0]).strip()) if r[0] else ""
        fval = str(r[5]).strip() if r[5] else ""
        if "מחוזי" in fval and "דיפרנציאלי" in fval and aval:
            norm = normalize_budget_name(aval)
            for b in _budgets_data:
                if b["norm_name"] == norm:
                    b["machozi_I"] += _to_num(r[8])

    # rcode → budget mapping from פירוט המענים (first occurrence wins)
    rcode_to_budget: dict = {}
    for r in perut_rows[1:]:
        braw = _norm_quotes(str(r[0]).strip()) if r[0] else ""
        rcode = str(r[17]).strip() if r[17] else ""
        if rcode and braw and rcode not in rcode_to_budget:
            rcode_to_budget[rcode] = normalize_budget_name(braw)

    # Per-budget yozma computation
    for bdata in _budgets_data:
        bname = bdata["norm_name"]
        budget_perut = [
            r for r in perut_rows[1:]
            if normalize_budget_name(_norm_quotes(str(r[0] or "").strip())) == bname
        ]
        # Deduplicate
        seen_b: dict = {}
        for r in budget_perut:
            key = tuple(str(x).strip() if x else "" for x in r[:10])
            if key not in seen_b:
                seen_b[key] = r
        unique_b = list(seen_b.values())

        plans_with_R_b = [r for r in unique_b if r[17] and str(r[17]).strip()]
        sum_chayav_b = sum(_to_num(r[15]) for r in plans_with_R_b)

        yb_code: dict = defaultdict(float)
        _comp_seen_b: set = set()
        for r in budget_perut:
            rcode = str(r[17] or "").strip()
            if rcode not in yozma_codes:
                continue
            _c13 = r[13] if len(r) > 13 else None
            if _c13 is not None:
                comp_key = (
                    str(r[9]  or "").strip(),
                    str(r[10] or "").strip(),
                    str(r[11] or "").strip(),
                    str(r[12] or "").strip(),
                    _c13,
                    rcode,
                )
            else:
                comp_key = (tuple(str(x).strip() if x else "" for x in r[:10]), rcode)
            if comp_key in _comp_seen_b:
                continue
            _comp_seen_b.add(comp_key)
            yb_code[rcode] += _to_num(_c13) if _c13 is not None else _to_num(r[15])
        total_yb = sum(yb_code.values())

        # Main גפן budget uses the globally-computed gamish (which deducts kvua gaps).
        # Other budgets have no kvua rows so H - L is the correct flexible amount.
        gamish_b = gamish_notar if bdata["norm_name"] == ctrl_norm else bdata["H"] - bdata["L"]
        max03_b = (bdata["H"] + bdata["machozi_I"]) * 0.3
        max04_b = (bdata["H"] + bdata["machozi_I"]) * 0.4

        bdata.update({
            "sum_chayav": sum_chayav_b,
            "plans_with_R_budget": plans_with_R_b,
            "gamish_notar": gamish_b,
            "yozma_03": _build_yozma_detail(yozma_codes, dict(yb_code), total_yb, max03_b, gamish_b),
            "yozma_04": _build_yozma_detail(yozma_codes, dict(yb_code), total_yb, max04_b, gamish_b),
        })

    return {
        "school_name": school_name,
        "school_code": int(school_code) if school_code else school_code,
        "school_stage": school_stage,
        "H": H,
        "L": L,
        "S": S,
        "T": T,
        "machozi_I": machozi_I,
        "abs_hefresh_kvua": abs_hefresh_kvua,
        "gamish_notar": gamish_notar,
        "sum_chayav": sum_chayav,
        "kvua_rows": kvua_rows,
        "has_multiple_budget_types": has_multiple_budget_types,
        "plans_with_R": plans_with_R,
        "yozma_codes": yozma_codes,
        "yozma_by_code": dict(yozma_by_code),
        "total_yozma_betikhnun": total_yozma_betikhnun,
        "yozma_03": yozma_03,
        "yozma_04": yozma_04,
        "rcode_to_budget": rcode_to_budget,
        "_budgets_data": _budgets_data,
    }


def cross_reference_doch(tikhnun_data, doch_filepath):
    """
    Cross-reference tikhnun plans with a Gefen doch execution file.
    Loads the raw doch file (not via gefen_processor, which filters rows).
    Adds execution sums, tanuz calculation, and partial rows to tikhnun_data.
    Returns updated tikhnun_data dict.
    """
    try:
        wb = openpyxl.load_workbook(doch_filepath, read_only=True)
        doch_rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
        wb.close()
    except Exception as e:
        raise ValueError(f"Failed to read doch file: {e}")

    exec_sums = defaultdict(float)
    nadche_sum = 0.0
    lelo_koved_sum = 0.0
    overlap_sum = 0.0

    for r in doch_rows[1:]:
        a_val = str(r[0]).strip() if r[0] else ""
        if "מענה משרדי" in a_val:
            continue

        l_val = _to_num(r[11])
        l_not_empty = bool(r[11]) and l_val != 0
        m_val = str(r[12]).strip() if r[12] else ""
        n_val = str(r[13]).strip() if r[13] else ""

        key = _plan_key_doch(r)
        exec_sums[key] += l_val

        is_nadche = m_val.startswith("נדחה:") and l_not_empty
        is_lelo = (n_val == "לא") and l_not_empty

        if is_nadche:
            nadche_sum += l_val
        if is_lelo:
            lelo_koved_sum += l_val
        if is_nadche and is_lelo:
            overlap_sum += l_val

    nikuy = nadche_sum + lelo_koved_sum - overlap_sum
    S = tikhnun_data["S"]
    sum_chayav = tikhnun_data["sum_chayav"]
    pct_tanuz = (S - nikuy) / sum_chayav if sum_chayav > 0 else 0

    # Cross-reference each plan with report code
    comparison_rows = []
    for r in tikhnun_data["plans_with_R"]:
        pkey = _plan_key_perut(r)
        tikhnun_val = _to_num(r[15])
        divuach = exec_sums.get(pkey, 0)
        hefresh = tikhnun_val - divuach
        pct = divuach / tikhnun_val if tikhnun_val > 0 else 0
        rcode = str(r[17]).strip() if r[17] else ""
        name = str(r[8]).strip() if r[8] else ""
        j_val = str(r[9]).strip() if r[9] else ""
        mispnum = j_val if (j_val and j_val != " ") else "אין"
        comparison_rows.append({
            "key": pkey,
            "rcode": rcode,
            "name": name,
            "mispnum": mispnum,
            "tikhnun": tikhnun_val,
            "divuach": divuach,
            "hefresh": hefresh,
            "pct": pct,
        })

    # Filter: only hefresh >= 1 (as specified)
    partial_rows = [x for x in comparison_rows if x["hefresh"] >= 1]
    partial_rows.sort(key=lambda x: (round(x["pct"], 4), -x["tikhnun"]))
    sum_hefresh_partial = sum(x["hefresh"] for x in partial_rows)

    # ── Yozma breakdown: per-initiative, per-supplier detail ─────────────────
    yozma_codes_set = set(tikhnun_data["yozma_codes"].keys())

    # initiative name lookup: (plan_number, code) → name
    plan_code_to_name = {}
    for r in tikhnun_data["plans_with_R"]:
        pn = str(r[9] or "").strip()
        cd = str(r[17] or "").strip()
        nm = str(r[8] or "").strip()
        if pn and cd:
            plan_code_to_name[(pn, cd)] = nm

    # group transactions: (plan_num, code) → supplier_num → [transaction]
    combo_sups: dict = defaultdict(lambda: defaultdict(list))
    for r in doch_rows[1:]:
        if str(r[0] or "").strip().startswith("מענה משרדי"):
            continue
        cd = str(r[1] or "").strip()
        if cd not in yozma_codes_set:
            continue
        pn = _extract_plan_num(str(r[3] or ""))
        if not pn:
            continue
        sup_raw = str(r[6] or "").strip()
        sup_m = re.match(r"^\s*(\d+)\s*-\s*", sup_raw)
        sup_num = sup_m.group(1) if sup_m else sup_raw
        sup_name = re.sub(r"^\s*\d+\s*-\s*", "", sup_raw).strip()
        try:
            amount = int(round(float(str(r[11] or "0").replace(",", "").strip()))) if r[11] else 0
        except Exception:
            amount = 0
        combo_sups[(pn, cd)][sup_num].append({
            "date": _fmt_date(r[5]),
            "invoice": str(r[4] or "").strip(),
            "description": str(r[10] or "").strip(),
            "amount": amount,
            "supplier_name": sup_name,
        })

    yozma_breakdown = []
    for (pn, cd), sup_dict in combo_sups.items():
        suppliers = []
        combo_total = 0.0
        for sup_num, txns in sup_dict.items():
            sup_total = sum(t["amount"] for t in txns)
            combo_total += sup_total
            suppliers.append({
                "supplier_number": sup_num,
                "supplier_name": txns[0]["supplier_name"],
                "total_amount": int(round(sup_total)),
                "transactions": [
                    {"date": t["date"], "invoice": t["invoice"],
                     "description": t["description"], "amount": t["amount"]}
                    for t in txns
                ],
            })
        suppliers.sort(key=lambda x: -x["total_amount"])
        yozma_breakdown.append({
            "plan_number": pn,
            "code": cd,
            "initiative_name": plan_code_to_name.get((pn, cd), ""),
            "total_amount": int(round(combo_total)),
            "suppliers": suppliers,
        })
    yozma_breakdown.sort(
        key=lambda x: (int(x["code"]) if x["code"].isdigit() else 0, x["plan_number"])
    )

    # ── Nihul breakdown: code 104 (tikkon) or 67 (beinayim) ──────────────────
    nihul_code = "104" if tikhnun_data.get("school_stage") == "תיכון" else "67"
    nihul_sups: dict = defaultdict(list)
    for r in doch_rows[1:]:
        if str(r[0] or "").strip().startswith("מענה משרדי"):
            continue
        cd = str(r[1] or "").strip()
        if cd != nihul_code:
            continue
        sup_raw = str(r[6] or "").strip()
        sup_m = re.match(r"^\s*(\d+)\s*-\s*", sup_raw)
        sup_num = sup_m.group(1) if sup_m else sup_raw
        sup_name = re.sub(r"^\s*\d+\s*-\s*", "", sup_raw).strip()
        try:
            amount = int(round(float(str(r[11] or "0").replace(",", "").strip()))) if r[11] else 0
        except Exception:
            amount = 0
        nihul_sups[sup_num].append({
            "date": _fmt_date(r[5]),
            "invoice": str(r[4] or "").strip(),
            "description": str(r[10] or "").strip(),
            "amount": amount,
            "supplier_name": sup_name,
        })

    nihul_suppliers = []
    nihul_total = 0.0
    for sup_num, txns in nihul_sups.items():
        sup_total = sum(t["amount"] for t in txns)
        nihul_total += sup_total
        nihul_suppliers.append({
            "supplier_number": sup_num,
            "supplier_name": txns[0]["supplier_name"],
            "total_amount": int(round(sup_total)),
            "transactions": [
                {"date": t["date"], "invoice": t["invoice"],
                 "description": t["description"], "amount": t["amount"]}
                for t in txns
            ],
        })
    nihul_suppliers.sort(key=lambda x: -x["total_amount"])
    nihul_breakdown = [{
        "code": nihul_code,
        "initiative_name": "ניהול ותפעול",
        "plan_number": "",
        "total_amount": int(round(nihul_total)),
        "suppliers": nihul_suppliers,
    }] if nihul_suppliers else []
    # ─────────────────────────────────────────────────────────────────────────

    # ── Distribute yozma_breakdown and nihul_breakdown per budget ─────────────
    _budgets_data = tikhnun_data.get("_budgets_data", [])
    rcode_to_budget = tikhnun_data.get("rcode_to_budget", {})
    if _budgets_data:
        default_bname = _budgets_data[0]["norm_name"]
        bname_to_idx = {b["norm_name"]: i for i, b in enumerate(_budgets_data)}

        for item in yozma_breakdown:
            bname = rcode_to_budget.get(item["code"], default_bname)
            idx = bname_to_idx.get(bname, 0)
            _budgets_data[idx]["yozma_breakdown"].append(item)

        # Mark the nihul-owning budget regardless of whether invoices exist.
        # This lets the frontend distinguish "no planning" from "planned but no invoices".
        nihul_owner = rcode_to_budget.get(nihul_code, default_bname)
        owner_idx = bname_to_idx.get(nihul_owner, 0)
        _budgets_data[owner_idx]["nihul_planned"] = True
        if nihul_breakdown:
            _budgets_data[owner_idx]["nihul_breakdown"] = nihul_breakdown
    # ─────────────────────────────────────────────────────────────────────────

    tikhnun_data.update({
        "has_doch": True,
        "pct_tanuz": pct_tanuz,
        "nikuy": nikuy,
        "partial_rows": partial_rows,
        "sum_hefresh_partial": sum_hefresh_partial,
        "yozma_breakdown": yozma_breakdown,
        "nihul_breakdown": nihul_breakdown,
    })
    return tikhnun_data


def build_tikhnun_result(tikhnun_data):
    """
    Build the serializable result dict for the API response.
    tikhnun_data may or may not have doch cross-reference.
    """
    kvua_has_issues = any(r["hefresh"] < 0 for r in tikhnun_data["kvua_rows"])
    partial_has_issues = bool(tikhnun_data.get("partial_rows"))

    overview = {
        "budget": tikhnun_data["H"],
        "planned": tikhnun_data["L"],
        "fixed_gap_abs": tikhnun_data["abs_hefresh_kvua"],
        "flexible_remaining": tikhnun_data["gamish_notar"],
        "sum_chayav": tikhnun_data["sum_chayav"],
        "sum_divuach": tikhnun_data["S"],
        "pct_divuach": tikhnun_data["T"],
        "pct_tanuz": tikhnun_data.get("pct_tanuz"),
    }

    # Build per-budget array
    budgets_out = []
    for b in tikhnun_data.get("_budgets_data", []):
        budgets_out.append({
            "name": b["norm_name"],
            "raw_name": b["raw_name"],
            "overview": {
                "budget": b["H"],
                "planned": b["L"],
                "sum_divuach": b["S"],
                "pct_divuach": b["T"],
                "flexible_remaining": b.get("gamish_notar", 0),
                "sum_chayav": b.get("sum_chayav", 0),
            },
            "yozma_03": b.get("yozma_03", {}),
            "yozma_04": b.get("yozma_04", {}),
            "yozma_breakdown": b.get("yozma_breakdown", []),
            "nihul_breakdown": b.get("nihul_breakdown", []),
            "nihul_planned": b.get("nihul_planned", False),
        })

    return {
        "school_name": tikhnun_data["school_name"],
        "school_code": tikhnun_data["school_code"],
        "school_stage": tikhnun_data["school_stage"],
        "filename": tikhnun_data.get("filename"),
        "has_doch": tikhnun_data.get("has_doch", False),
        "overview": overview,
        "kvua_rows": tikhnun_data["kvua_rows"],
        "kvua_has_issues": kvua_has_issues,
        "has_multiple_budget_types": tikhnun_data["has_multiple_budget_types"],
        "partial_rows": tikhnun_data.get("partial_rows", []),
        "partial_has_issues": partial_has_issues,
        "sum_hefresh_partial": tikhnun_data.get("sum_hefresh_partial", 0),
        # top-level yozma/nihul kept for backward compat (dialog, etc.)
        "yozma_03": tikhnun_data["yozma_03"],
        "yozma_04": tikhnun_data["yozma_04"],
        "yozma_breakdown": tikhnun_data.get("yozma_breakdown", []),
        "nihul_breakdown": tikhnun_data.get("nihul_breakdown", []),
        # per-budget array
        "budgets": budgets_out,
    }
