"""Combined PDF and Excel export across multiple sections."""

from io import BytesIO

import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, SimpleDocTemplate

from logic.pdf_exporter import (
    _ensure_fonts,
    build_hashva_section_story,
    build_school_info_story,
)
from logic.tikhnun_exporter import (
    _write_kvua,
    _write_partial,
    _write_sikar,
    _write_yozma,
    build_tikhnun_section_story,
)

TIKHNUN_TABS = {"sikar", "kvua", "partial", "yozma"}
HASHVA_TABS = {"hashva", "rejected", "nopdf"}

_SHEET_NAMES = {
    "sikar":   "סקירה - גפן",
    "kvua":    "מימוש תקציב קבוע",
    "partial": "תוכניות עם ביצוע חלקי",
    "yozma":   "יוזמות וצרכים",
}


def export_combined_pdf(run_data: dict, sections: list, multiplier: str = "03") -> bytes:
    _ensure_fonts()
    tikhnun = (run_data or {}).get("tikhnun")
    has_tikhnun = bool(tikhnun and not tikhnun.get("error"))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm,
    )
    story = []

    if has_tikhnun:
        story.extend(build_school_info_story(tikhnun))
        if sections:
            story.append(PageBreak())

    for i, section in enumerate(sections):
        if i > 0:
            story.append(PageBreak())
        if section in TIKHNUN_TABS and has_tikhnun:
            story.extend(build_tikhnun_section_story(tikhnun, section, multiplier))
        elif section in HASHVA_TABS:
            story.extend(build_hashva_section_story(run_data or {}, section))

    if not story:
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import Paragraph
        story = [Paragraph("אין נתונים", ParagraphStyle("empty"))]

    doc.build(story)
    return buf.getvalue()


def export_combined_excel(run_data: dict, sections: list, multiplier: str = "03") -> bytes:
    tikhnun = (run_data or {}).get("tikhnun")
    has_tikhnun = bool(tikhnun and not tikhnun.get("error"))

    hashva_sections = [s for s in sections if s in HASHVA_TABS]
    tikhnun_sections = [s for s in sections if s in TIKHNUN_TABS]

    if hashva_sections and (run_data or {}).get("file_path"):
        wb = openpyxl.load_workbook(run_data["file_path"])
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if has_tikhnun and tikhnun_sections:
        yozma_key = "yozma_04" if multiplier == "04" else "yozma_03"
        write_fns = {"sikar": _write_sikar, "kvua": _write_kvua, "partial": _write_partial}
        for section in tikhnun_sections:
            base_title = _SHEET_NAMES.get(section, section)
            title = base_title
            n = 1
            while title in wb.sheetnames:
                n += 1
                title = f"{base_title} ({n})"
            ws = wb.create_sheet(title=title)
            ws.sheet_view.rightToLeft = True
            if section == "yozma":
                _write_yozma(ws, tikhnun, yozma_key)
            elif section in write_fns:
                write_fns[section](ws, tikhnun)

    if not wb.sheetnames:
        wb.create_sheet("נתונים")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
